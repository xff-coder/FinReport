import freecashflow
import fixed_assets
import statementFunct
import styleModule
import numpy as np
import requests
import json
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle, Border, Side, Color
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE
from openpyxl.utils import get_column_letter
from string import ascii_uppercase
        
def increment_letter(letter, increment):
    # Function to increment a letter by a specified number
    if increment == 0:
        return letter
    else:
        alpha_index = ascii_uppercase.index(letter.upper()) + increment
        return ascii_uppercase[alpha_index % 26] + "A" * (alpha_index // 26)
    
# Define the categories for the Net Working Capital sheet from the image provided
categories = [
        "Cash & Cash Equivalents",
        "Accounts Receivable",
        "Inventory",
        "Other Current Assets",
        "Total Current Assets",
        " ",  # Blank row for spacing
        "Accounts Payable",
        "Other Payables & Accruals",
        "Short Term Debt",
        "Other Short Term Liabilities",
        "Total Current Liabilities",
        "", "", "", # Multiple blank rows to separate the tables visually
        "Assumptions",
        "Fiscal Year",
        "Revenue",
        "COGS",
        " ",
        "Days Sales Outstanding (DSO)",
        "Days Inventory Outstanding (DIO)",
        "Days Payable Outstanding (DPO)",
        " ",
        "Cash & Cash Equivalents Ratio",
        "Other Current Assets as a % of Revenue",
        "Short Term Debt as a % of Revenue",
        "Other Short Term Liabilities as a % of Revenue",
        "Other Payables & Accruals as a % of Revenue"
]

def apply_bold_to_specific_rows(worksheet, statementName):
    if statementName == 'Net Working Capital':
        bold_rows = [
            "Total Current Assets","Total Current Liabilities"
        ]
    
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):  # Starting from row 2
        if row[0].value in bold_rows:
            for cell in row:
                cell.font = Font(bold=True)

            # Apply thick border to the top of the entire row after the bolded row
            next_row_idx = row[0].row
            if next_row_idx <= worksheet.max_row:
                for next_row_cell in worksheet[next_row_idx]:
                    next_row_cell.border = Border(top=Side(style='thin'))

    # Apply bold formatting only to specific cells in Column A
    for cell in worksheet['A']:  # Iterate over all cells in Column A
        if cell.value in bold_rows:
            cell.font = Font(bold=True)
        else:
            cell.font = Font(bold=False)
   
   

def style_asset_row(worksheet, color_hex):
    # Define medium border style
    medium_border = Border(top=Side(style='medium'), bottom=Side(style='medium'))
    fill_color = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

    for row in worksheet.iter_rows(min_row=1, max_col=1):
        for cell in row:
            if cell.value == "Total Current Assets":
                row_number = cell.row
                # Apply a medium border above and below the row, and fill color to all cells in this row
                for col in range(1, worksheet.max_column + 1):
                    cell_to_style = worksheet.cell(row=row_number, column=col)
                    cell_to_style.border = medium_border
                    cell_to_style.fill = fill_color

def style_liabilities_row(worksheet, color_hex):
    # Define medium border style
    medium_border = Border(top=Side(style='medium'), bottom=Side(style='medium'))
    fill_color = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

    # Find the row number for "Unlevered Free Cash Flow"
    for row in worksheet.iter_rows(min_row=1, max_col=1):
        for cell in row:
            if cell.value == "Total Current Liabilities":
                row_number = cell.row
                # Apply a medium border above and below the row, and fill color to all cells in this row
                for col in range(1, worksheet.max_column + 1):
                    cell_to_style = worksheet.cell(row=row_number, column=col)
                    cell_to_style.border = medium_border
                    cell_to_style.fill = fill_color
                    
def apply_percentage_format_to_rows(worksheet):
    # Rows to be formatted as percentages
    percentage_rows = [
        "Cash & Cash Equivalents Ratio",
        "Other Current Assets as a % of Revenue",
        "Short Term Debt as a % of Revenue",
        "Other Short Term Liabilities as a % of Revenue",
        "Other Payables & Accruals as a % of Revenue"    ]

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):  # Starting from row 2
        if row[0].value in percentage_rows:
            for cell in row:
                cell.number_format = '0.00%'
                                
def AddNetWorkingCapitalSheet(start_year, end_year):
    # Generate fiscal years starting from start_year to end_year inclusive
    fiscal_years = list(range(start_year, end_year + 1)) + [f"{year}E" for year in range(end_year + 1, end_year + 6)]
    fiscal_years = list(map(str, fiscal_years))
    
    data = {"Category": categories}
    dso_formulas = []  
    dio_formulas = []
    dpo_formulas = []
    cce_formulas = []
    oca_percent_rev_formulas = []
    std_percent_rev_formulas = []
    ostl_percent_rev_formulas = []
    opa_percent_rev_formulas = []

    for i, year in enumerate(fiscal_years):
        is_estimated_year = year.endswith("E")
        data[year] = []
        for category in categories:
            col_letter = increment_letter('B', i)
            if category == "Accounts Receivable":
                if is_estimated_year:
                    current_year_revenue_cell = f"{col_letter}19"
                    ar_percent_revenue_cell = f"{col_letter}22"
                    formula = f"={current_year_revenue_cell} * {ar_percent_revenue_cell} / 365"
                else:
                    formula = f"='Balance Sheet'!{col_letter}7"
                    
            elif category == "Cash & Cash Equivalents":
                if is_estimated_year:
                    current_year_revenue_cell = f"{col_letter}19"
                    cce_percent_revenue_cell = f"{col_letter}26"
                    formula = f"={current_year_revenue_cell} * {cce_percent_revenue_cell}"
                else:
                    formula = f"='Balance Sheet'!{col_letter}3"
                    
            elif category == "Inventory":
                if is_estimated_year:
                    current_year_revenue_cell = f"{col_letter}19"
                    inv_percent_revenue_cell = f"{col_letter}23"
                    formula = f"={current_year_revenue_cell} * {inv_percent_revenue_cell} / 365"
                else:
                    formula = f"='Balance Sheet'!{col_letter}8"
            elif category == "Other Current Assets":
                if is_estimated_year:
                    current_year_revenue_cell = f"{col_letter}19"
                    oca_percent_revenue_cell = f"{col_letter}27"
                    formula = f"={current_year_revenue_cell} * {oca_percent_revenue_cell}"
                else:
                    formula = f"='Balance Sheet'!{col_letter}9"
                    
            elif category == "Total Current Assets":
                formula = f"=SUM({col_letter}3:{col_letter}6)"
                
            elif category == "Accounts Payable":
                if is_estimated_year:
                    current_year_revenue_cell = f"{col_letter}19"
                    ap_percent_revenue_cell = f"{col_letter}23"
                    formula = f"={current_year_revenue_cell} * {ap_percent_revenue_cell} / 365"
                else:
                    formula = f"='Balance Sheet'!{col_letter}22"
                    
            elif category == "Other Payables & Accruals":
                if is_estimated_year:
                    current_year_revenue_cell = f"{col_letter}19"
                    opa_percent_revenue_cell = f"{col_letter}30"
                    formula = f"={current_year_revenue_cell} * {opa_percent_revenue_cell}"
                else:
                    formula = f"='Balance Sheet'!{col_letter}23"
                    
            elif category == "Short Term Debt":
                if is_estimated_year:
                    current_year_revenue_cell = f"{col_letter}19"
                    std_percent_revenue_cell = f"{col_letter}28"
                    formula = f"={current_year_revenue_cell} * {std_percent_revenue_cell}"
                else:
                    formula = f"='Balance Sheet'!{col_letter}24"
                    
            elif category == "Other Short Term Liabilities":
                if is_estimated_year:
                    current_year_revenue_cell = f"{col_letter}19"
                    ostl_percent_revenue_cell = f"{col_letter}29"
                    formula = f"={current_year_revenue_cell} * {ostl_percent_revenue_cell}"
                else:
                    formula = f"='Balance Sheet'!{col_letter}25"
                    
            elif category == "Total Current Liabilities":
                formula = f"=SUM({col_letter}9:{col_letter}12)"
                
            elif category == "Revenue":
                if is_estimated_year:
                    prev_col_letter = increment_letter('B', i - 1)
                    formula = f"='Free Cash Flow'!{prev_col_letter}3"
                else:
                    formula = f"='Income Statement'!{col_letter}3"
            elif category == "COGS":
                if is_estimated_year:
                    prev_col_letter = increment_letter('B', i - 1)
                    formula = f"='Free Cash Flow'!{prev_col_letter}4"
                else:
                    formula = f"='Income Statement'!{col_letter}4"
            elif category == "Days Sales Outstanding (DSO)":
                if is_estimated_year:
                    if dso_formulas:
                        avg_formula = f"=AVERAGE({','.join(dso_formulas)})"
                        formula = avg_formula
                    else:
                        formula = None  # If no previous data, set to None or an appropriate placeholder
                else:
                    dso_formula = f"={col_letter}4 / {col_letter}19 * 365"  # Assuming DSO formula structure
                    dso_formulas.append(f"{col_letter}{categories.index('Days Sales Outstanding (DSO)') + 3}")
                    formula = dso_formula            
            
            elif category == "Days Inventory Outstanding (DIO)":
                if is_estimated_year:
                    if dio_formulas:
                        avg_formula = f"=AVERAGE({','.join(dio_formulas)})"
                        formula = avg_formula
                    else:
                        formula = None 
                else:
                    dio_formula = f"={col_letter}5 / {col_letter}19 * 365" 
                    dio_formulas.append(f"{col_letter}{categories.index('Days Inventory Outstanding (DIO)') + 3}")
                    formula = dio_formula   
            elif category == "Days Payable Outstanding (DPO)":
                if is_estimated_year:
                    if dpo_formulas:
                        avg_formula = f"=AVERAGE({','.join(dpo_formulas)})"
                        formula = avg_formula
                    else:
                        formula = None
                else:
                    dpo_formula = f"={col_letter}9 / {col_letter}19 * 365" 
                    dpo_formulas.append(f"{col_letter}{categories.index('Days Payable Outstanding (DPO)') + 3}")
                    formula = dpo_formula
                    
            elif category == "Cash & Cash Equivalents Ratio":
                if is_estimated_year:
                    if cce_formulas:
                        avg_formula = f"=AVERAGE({','.join(cce_formulas)})"
                        formula = avg_formula
                    else:
                        formula = None
                else:
                    cce_formula = f"={col_letter}3 / {col_letter}19" 
                    cce_formulas.append(f"{col_letter}{categories.index('Cash & Cash Equivalents Ratio') + 3}")
                    formula = cce_formula   
                
            elif category == "Other Current Assets as a % of Revenue":
                if is_estimated_year:
                    if oca_percent_rev_formulas:
                        avg_formula = f"=AVERAGE({','.join(oca_percent_rev_formulas)})"
                        formula = avg_formula
                    else:
                        formula = None
                else:
                    oca_percent_rev_formula = f"={col_letter}6 / {col_letter}19" 
                    oca_percent_rev_formulas.append(f"{col_letter}{categories.index('Other Current Assets as a % of Revenue') + 3}")
                    formula = oca_percent_rev_formula  
                        
            elif category == "Short Term Debt as a % of Revenue":
                if is_estimated_year:
                    if std_percent_rev_formulas:
                        avg_formula = f"=AVERAGE({','.join(std_percent_rev_formulas)})"
                        formula = avg_formula
                    else:
                        formula = None
                else:
                    std_percent_rev_formula = f"={col_letter}11 / {col_letter}19"  
                    std_percent_rev_formulas.append(f"{col_letter}{categories.index('Short Term Debt as a % of Revenue') + 3}")
                    formula = std_percent_rev_formula
                    
            elif category == "Other Short Term Liabilities as a % of Revenue":
                if is_estimated_year:
                    if ostl_percent_rev_formulas:
                        avg_formula = f"=AVERAGE({','.join(ostl_percent_rev_formulas)})"
                        formula = avg_formula
                    else:
                        formula = None
                else:
                    ostl_percent_rev_formula = f"={col_letter}12 / {col_letter}19"  
                    ostl_percent_rev_formulas.append(f"{col_letter}{categories.index('Other Short Term Liabilities as a % of Revenue') + 3}")
                    formula = ostl_percent_rev_formula
            elif category == "Other Payables & Accruals as a % of Revenue":
                if is_estimated_year:
                    if opa_percent_rev_formulas:
                        avg_formula = f"=AVERAGE({','.join(opa_percent_rev_formulas)})"
                        formula = avg_formula
                    else:
                        formula = None
                else:
                    opa_percent_rev_formula = f"={col_letter}10 / {col_letter}19"  
                    opa_percent_rev_formulas.append(f"{col_letter}{categories.index('Other Payables & Accruals as a % of Revenue') + 3}")
                    formula = opa_percent_rev_formula
            else:    
                formula = " "  # Placeholder for other categories
            data[year].append(formula)

            

                    
                    
    df = pd.DataFrame(data)
    df.set_index('Category', inplace=True)
    return df