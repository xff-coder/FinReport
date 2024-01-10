import freecashflow
import fixed_assets
import statementFunct
import styleModule
import numpy as np
import requests
import json
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle, Border, Side, Color
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE
from openpyxl.utils import get_column_letter
from string import ascii_uppercase

def increment_letter(letter, increment):
    # Function to increment a letter by a specified number
    if increment == 0:
        return letter
    else:
        alpha_index = ascii_uppercase.index(letter.upper()) + increment
        return ascii_uppercase[alpha_index % 26] + "A" * (alpha_index // 26)


def style_ending_ppe_row(worksheet, color_hex):
    # Define medium border style
    medium_border = Border(top=Side(style='medium'), bottom=Side(style='medium'))
    fill_color = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

    for row in worksheet.iter_rows(min_row=1, max_col=1):
        for cell in row:
            if cell.value == "Ending PP&E":
                row_number = cell.row
                for col in range(1, worksheet.max_column + 1):
                    cell_to_style = worksheet.cell(row=row_number, column=col)
                    cell_to_style.border = medium_border
                    cell_to_style.fill = fill_color
                    
                    
def auto_adjust_column_width(worksheet, max_width=15, first_col_min_width=15):
    for column in worksheet.columns:
        max_length = 0
        col_letter = column[0].column_letter  # Get the column letter

        for cell in column:
            if cell.value:
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)

        if col_letter == 'A':  # Special handling for the first column
            adjusted_width = max(max_length + 2, first_col_min_width)
        else:
            # Apply a tighter limit for non-column A columns
            adjusted_width = min(max_length, max_width)

        worksheet.column_dimensions[col_letter].width = adjusted_width


def find_last_non_estimated_ending_PPE_cell(start_year, end_year):
    # Determine the last non-estimated fiscal year
    last_non_estimated_year = end_year

    # Calculate the number of increments needed from the base column ('C' in this case)
    increments = last_non_estimated_year - (start_year + 1)

    # Find the column letter
    col_letter = increment_letter('C', increments)

    # Assuming "Ending PP&E" is in a specific row (e.g., row 12)
    row_number = 12

    # Construct the cell reference
    cell_reference = f"{col_letter}{row_number}"
    return cell_reference

def apply_percentage_format_to_rows(worksheet):
    # Rows to be formatted as percentages
    percentage_rows = [
"D&A as a % of Beginning PP&E", "CapEx as a % of Beginning PP&E"
    ]

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):  # Starting from row 2
        if row[0].value in percentage_rows:
            for cell in row:
                cell.number_format = '0.00%'
                
def AddSheet(start_year, end_year):
    # Adjust the start year for the fixed assets sheet
    adjusted_start_year = start_year + 1
    row_number_for_DA_percentage = 10
    row_number_for_CapEx_percentage = 11

    # Generate fiscal years starting from adjusted_start_year
    fiscal_years = list(range(adjusted_start_year, end_year + 1)) + [f"{year}E" for year in range(end_year + 1, end_year + 6)]
    fiscal_years = list(map(str, fiscal_years))

    categories = ["Beginning PP&E", "D&A", "CapEx", "Ending PP&E",
                  " ", "Assumptions", "Fiscal Year",
                  "D&A as a % of Beginning PP&E", "CapEx as a % of Beginning PP&E"]

    data = {"Category": categories}
    
    
    for i, year in enumerate(fiscal_years):
        is_estimated_year = year.endswith("E")
        data[year] = []
        
        col_letter = increment_letter('B', i)  # Define col_letter for the current year
        next_col_letter = increment_letter(col_letter, 1)  # Shift one column ahead
        for category in categories:
            col_letter = increment_letter('B', i)
            if category == "Beginning PP&E":
                formula = f"='Balance Sheet'!{col_letter}12"
            elif category == "D&A":
                formula = f"='Cash Flow'!{next_col_letter}4"
            elif category == "CapEx":
                formula = f"={col_letter}6 - {col_letter}3 + {col_letter}4"
            elif category == "Ending PP&E":
                col_letter = increment_letter('C', i) 
                formula = f"='Balance Sheet'!{next_col_letter}12"
            elif category == "D&A as a % of Beginning PP&E":
                formula = f"={col_letter}4 / {col_letter}3"
            elif category == "CapEx as a % of Beginning PP&E":
                formula = f"={col_letter}5 / {col_letter}3"
            else:
                formula = " "  # Placeholder for other categories
            data[year].append(formula)
            
    # Additional logic for estimated fiscal years
    for i, year in enumerate(fiscal_years):
        if year.endswith("E"):  # Focus only on estimated years
            col_letter = increment_letter('B', i)  # Column for the current year
            prev_year_col_letter = increment_letter('C', i - 2)  # Column for the previous year's "Ending PP&E"

            for category in categories:
                if category == "Beginning PP&E":
                    # Reference the 'Ending PP&E' from the previous fiscal year in the same sheet
                    previous_year_ending_ppe_cell = f"{prev_year_col_letter}6"
                    data[year][categories.index(category)] = f"={previous_year_ending_ppe_cell}"

    # Calculate the average for 'D&A as a % of Beginning PP&E' for E fiscal years
    da_percentage_cells = []
    for i, year in enumerate(fiscal_years):
        if year.endswith("E"):
            col_letter = increment_letter('B', i - 5)
            da_percentage_cell = f"{col_letter}{row_number_for_DA_percentage}"
            da_percentage_cells.append(da_percentage_cell)

    # Formula to calculate the average
    if da_percentage_cells:
        average_formula = f"=AVERAGE({','.join(da_percentage_cells)})"

        # Apply this average to each 'E' fiscal year in 'D&A as a % of Beginning PP&E'
        for i, year in enumerate(fiscal_years):
            if year.endswith("E"):
                data[year][categories.index("D&A as a % of Beginning PP&E")] = average_formula

    # Logic for 'CapEx as a % of Beginning PP&E'
    capex_percentage_cells = []
    for i, year in enumerate(fiscal_years):
        if year.endswith("E"):
            col_letter = increment_letter('B', i - 5)
            capex_percentage_cell = f"{col_letter}{row_number_for_CapEx_percentage}"
            capex_percentage_cells.append(capex_percentage_cell)

    # Formula to calculate the average for CapEx
    if capex_percentage_cells:
        average_formula_capex = f"=AVERAGE({','.join(capex_percentage_cells)})"

        # Apply this average to each 'E' fiscal year in 'CapEx as a % of Beginning PP&E'
        for i, year in enumerate(fiscal_years):
            if year.endswith("E"):
                data[year][categories.index("CapEx as a % of Beginning PP&E")] = average_formula_capex
                
    # Logic for the D&A 'E' formula
    for i, year in enumerate(fiscal_years):
        if year.endswith("E"):  # Target only the 'E' fiscal years
            col_letter = increment_letter('B', i)  # Column for the current 'E' fiscal year
            value_in_row_3 = f"{col_letter}3"
            value_in_row_9 = f"{col_letter}10"

            # Formula to multiply the values in row 3 and row 7
            multiplication_formula = f"={value_in_row_3} * {value_in_row_9}"

            # Assign the formula to the 'D&A' category for the current 'E' fiscal year
            data[year][categories.index("D&A")] = multiplication_formula
            
    # Logic for the CapEx 'E' formula
    for i, year in enumerate(fiscal_years):
        if year.endswith("E"):  # Target only the 'E' fiscal years
            col_letter = increment_letter('B', i)  # Column for the current 'E' fiscal year
            value_in_row_3 = f"{col_letter}3"
            value_in_row_10 = f"{col_letter}11"

            # Formula to multiply the values in row 3 and row 8
            multiplication_formula = f"={value_in_row_3} * {value_in_row_10}"

            # Assign the formula to the 'CapEx' category for the current 'E' fiscal year
            data[year][categories.index("CapEx")] = multiplication_formula
            
    # Logic for the Ending PP&E 'E' formula
    for i, year in enumerate(fiscal_years):
        if year.endswith("E"):  # Target only the 'E' fiscal years
            col_letter = increment_letter('B', i)  # Column for the current 'E' fiscal year
            value_in_row_3 = f"{col_letter}3"
            value_in_row_4 = f"{col_letter}4"
            value_in_row_5 = f"{col_letter}5"

            # Formula for Ending PP&E: (Year 'E' Row 3) - (Year 'E' Row 4) + (Year 'E' Row 5)
            formula_ending_ppe = f"={value_in_row_3} - {value_in_row_4} + {value_in_row_5}"

            # Assign the formula to the 'Ending PP&E' category for the current 'E' fiscal year
            data[year][categories.index("Ending PP&E")] = formula_ending_ppe

                    
    df = pd.DataFrame(data)
    df.set_index('Category', inplace=True)
    return df