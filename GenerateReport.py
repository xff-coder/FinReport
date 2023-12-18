import FinModule 
import numpy as np
import requests
import json
import pandas as pd
import os
import sys
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle, Border, Side, Color
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
from openpyxl.utils import get_column_letter
from datetime import datetime
import nwc
import freecashflow
import styleModule
import fixed_assets
import statementFunct

# Define the currency format style
currency_style = NamedStyle(name='currency_style')
currency_style.number_format = FORMAT_CURRENCY_USD_SIMPLE
currency_style.alignment = Alignment(horizontal='right')

# Define a style for fiscal year
fiscal_year_style = NamedStyle(name='fiscal_year_style')
fiscal_year_style.font = Font(color=Color(rgb='002596BE'), bold=True) # RGB for the desired blue color
fiscal_year_style.alignment = Alignment(horizontal='center', vertical='center') # Center alignment
# Define a thick top border style
thick_top_border = Border(top=Side(style='thick'))


def AddMoreSheet():
    # Define the named styles for the workbook
    # Add the named styles to the workbook
    
    # Replace the lists with the actual data from the image
    columns = ["Fiscal Year", "2018", "2019", "2020", "2021", "2022", "2023E", "2024E", "2025E", "2026E", "2027E"]

    fixed_assets_data = [
    ["Fixed Assets Schedule"],
    ["Fiscal Year", "2018", "2019", "2020", "2021", "2022", "2023E", "2024E", "2025E", "2026E", "2027E"],
    ["Beginning PP&E", "='Balance Sheet'!C12", "...", "...", "...", "...", "...", "...", "...", "...", "..."],
    ["D&A", "...", "...", "...", "...", "...", "...", "...", "...", "...", "..."],
    ["CapEx", "=PPE-DA+CAPEX", "...", "...", "...", "...", "...", "...", "...", "...", "..."],
    ["Ending PP&E", "...", "...", "...", "...", "...", "...", "...", "..."],
    [],
    ["Assumptions"],
    ["Fiscal Year", "2018", "2019", "2020", "2021", "2022", "2023E", "2024E", "2025E", "2026E", "2027E"],
    ["D&A as a % of Beginning PP&E", "...", "...", "...", "...", "...", "...", "...", "..."],
    ["CapEx as a % of Beginning PP&E", "...", "...", "...", "...", "...", "...", "...", "..."]
    ]
 
    df = pd.DataFrame(fixed_assets_data, columns=columns)
    df.set_index('Fiscal Year', inplace=True)
    return df.T

    # Write the fixed assets data

def CheckForFile(reportRoot, curdate, filename):
    
    full_path = os.path.join(reportRoot + curdate, filename)

    # Check if the folder {curdate} exists, if not, create it
    if not os.path.exists(reportRoot + curdate):
        os.makedirs(reportRoot + curdate)
    elif os.path.exists(full_path):
        return True
    return False      
  
def main():
    #Main Code
    # Prompt the user for a ticker symbol
    # ticker_symbol = input("Please enter the ticker symbol: ")
    # start_year = input("Enter the start fiscal year (e.g., 2017): ")
    # end_year = input("Enter the end fiscal year (e.g., 2022): ")

    # # Convert start_year and end_year to integers
    # start_year = int(start_year)
    # end_year = int(end_year)

    # Generate a list of years from start_year to end_year
    ticker_symbol = sys.argv[2]
    start_year = int(sys.argv[3])
    end_year = int(sys.argv[4])
    statement_types = ','.join(sys.argv[5:])
    reportRoot = 'reports/'
    os.chdir (reportRoot);
    years = list(range(start_year, end_year + 1))
    years_str = ','.join(map(str, years)) # Convert the list of years to a comma-separated string
    curdate = datetime.now().strftime('%Y%m%d')
    filename = ticker_symbol + '_' + '_'.join(sys.argv[5:]) + '_' + "_".join(map(str, years)) + '.xlsx'
  
    if CheckForFile(reportRoot, curdate, filename) == True:
        print(curdate + '/' + filename)
        return

    filename = os.path.join(curdate, filename)
    # Construct the full path with /{curdate}/{filename}
    # Construct the URL with the user's input
    url = f"https://backend.simfin.com/api/v3/companies/statements/compact?ticker={ticker_symbol}&statements={statement_types}&period=FY&fyear={years_str}"

    headers = {
    "accept": "application/json",
    "Authorization": "b1c124d9-f078-4887-abb9-d3504b54b23b"
    }

    response = requests.get(url, headers=headers)
    jsonStr = json.loads(response.text)

    data = jsonStr[0]['statements']
    arrDataFrames = np.empty(len(data), dtype=object)
    arrNames = np.empty(len(data), dtype=object)

    for i in range(len(data)):
        arrNames[i] = FinModule.getStatementName(data[i]['statement'])
        columns = data[i]['columns']
        # Extract data
        table_data = data[i]['data']
        neededColumns = FinModule.getColumns(data[i]['statement'])
        # Create a DataFrame
        arrDataFrames[i] = FinModule.process_statement(table_data, columns, neededColumns) 
    newSheet = AddMoreSheet()
    
    fixed_assets_Sheet = fixed_assets.AddSheet(start_year, end_year)
    free_cash_flow_sheet = freecashflow.AddFreeCashFlowSheet(start_year, end_year)
    nwc_sheet = nwc.AddNetWorkingCapitalSheet(start_year, end_year)

    # Write the filtered data to an Excel file with formatting
    with pd.ExcelWriter(reportRoot + filename, engine='openpyxl') as writer:
    # Register the named styles with the writer's workbook
        writer.book.add_named_style(currency_style)
        writer.book.add_named_style(fiscal_year_style)
        for i in range(len(data)):
        # Write the DataFrame to the worksheet
            arrDataFrames[i].to_excel(writer, sheet_name=arrNames[i], startrow=1)

            # Get the worksheet object
            worksheet = writer.sheets[arrNames[i]]
            # Apply bold formatting to specific rows
            statementFunct.apply_bold_to_specific_rows(worksheet, arrNames[i])
            styleModule.SetOtherStyle(worksheet)
        for sheet_name in writer.book.sheetnames:
            statementFunct.auto_adjust_column_width(writer.book[sheet_name])
        newSheet.to_excel(writer, sheet_name='definition', startrow=1)
        fixed_assets.CreateSheet(writer, fixed_assets_Sheet, "Fixed Assets")
        freecashflow.CreateSheet(writer, free_cash_flow_sheet, "Free Cash Flow", start_year, end_year)
        nwc.CreateSheet(writer, nwc_sheet, "Net Working Capital")
    
    print(filename)

if __name__ == "__main__":
    main()