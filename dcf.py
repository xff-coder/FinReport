import nwc
import wacc
import freecashflow
import fixed_assets
import statementFunct
import styleModule
import numpy as np
import requests
import json
import pandas as pd
import openpyxl
import yfinance as yf
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle, Border, Side, Color
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell import MergedCell
from string import ascii_uppercase


def increment_letter(letter, increment):
    # Function to increment a letter by a specified number
    if increment == 0:
        return letter
    else:
        alpha_index = ascii_uppercase.index(letter.upper()) + increment
        return ascii_uppercase[alpha_index % 26] + "A" * (alpha_index // 26)
    
# Define the categories for the Net Working Capital sheet from the image provided
def get_shares_outstanding(ticker):
    stock = yf.Ticker(ticker)
    info = stock.info
    shares_outstanding = info.get('sharesOutstanding')
    return shares_outstanding


def create_centered_sensitivity_table(current_wacc, current_growth, increment=0.5, steps=4):

    current_wacc = f"= B11" 
    current_growth = f"= B10"

    wacc_range = [(current_wacc + (i - steps) * increment) for i in range(2 * steps + 1)]
    growth_range = [(current_growth + (i - steps) * increment) for i in range(2 * steps + 1)]

    # Create a DataFrame to hold the sensitivity table
    data = {'Growth Rate \\ WACC': wacc_range}
    for growth in growth_range:
        data[f"{growth}%"] = [f"=YourCalculationFunction({wacc}, {growth})" for wacc in wacc_range]

    df = pd.DataFrame(data)
    df.set_index('Growth Rate \\ WACC', inplace=True)
    return df

                

def format_rows(worksheet):
    # Define border styles
    thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    medium_border = Border(top=Side(style='medium'), bottom=Side(style='medium'))
    isp_fill = PatternFill(start_color='fdfd96', end_color='fdfd96', fill_type='solid')

    # Categories to be bolded and have thin border
    bold_and_thin_border_categories = [
        "Present Value of Free Cash Flow"
    ]

    # Category to be bolded and have medium border
    bold_and_medium_border_category = ["Implied Share Price"]

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
        category = row[0].value
        if category in bold_and_thin_border_categories:
            # Apply bold font and thin border
            for cell in row:
                cell.font = Font(bold=True)
                cell.border = thin_border

        elif category in bold_and_medium_border_category:
            # Apply bold font, medium border, and fill for 'Implied Share Price' up to column 2
            for cell in row[:2]:  # Limit to first two columns
                cell.font = Font(bold=True)
                cell.border = medium_border
                cell.fill = isp_fill
                


    
def add_dcf_sheet(start_year, end_year, ticker_symbol):
    
    total_shares = get_shares_outstanding(ticker_symbol) / 1000000  # Convert to millions if necessary

           

    # Separate the categories into two groups
    categories_group_1 = [
        "Unlevered Free Cash Flow",
        " ",
        "Projection Year",
        "Present Value of Free Cash Flow"
    ]

    categories_group_2 = [
        "Sum of PV of FCF",
        "Growth Rate",
        "WACC",
        "Terminal Value",
        "PV of Terminal Value",
        "Enterprise Value",
        "(+) Cash",
        "(-) Debt",
        "(-) Minority Interest",
        "Equity Value",
        "Total Shares Outstanding (mm)",
        "Implied Share Price"
    ]

    # Generate fiscal years
    fiscal_years = list(range(start_year + 1, end_year + 1)) + [f"{year}E" for year in range(end_year + 1, end_year + 6)]
    fiscal_years = list(map(str, fiscal_years))

    # Identify the column letter for the last non-E fiscal year
    last_fiscal_year_index = len([year for year in fiscal_years if not year.endswith("E")])
    last_fiscal_year_col_letter = increment_letter('B', last_fiscal_year_index)
    last_non_e_fiscal_year_index = len([year for year in fiscal_years if not year.endswith("E")]) - 1
    last_e_fiscal_year_index = len(fiscal_years) - 1
    last_non_e_fiscal_year_col_letter = increment_letter('B', last_non_e_fiscal_year_index)
    last_e_fiscal_year_col_letter = increment_letter('B', last_e_fiscal_year_index)
    
    data = {"Category": categories_group_1 + [" "] + [" "] + categories_group_2}
    for year in fiscal_years:
        data[year] = [None] * len(data["Category"])

    projection_counter = 1
    sum_formula_parts = []  # To hold the cell references for summing
    pvsum_formula_parts = []
    last_pvfcf_cell = None  # To hold the cell reference for the last E year's UFCF
    last_ufcf_cell = None  # To hold the cell reference for the last E year's UFCF

    # Populate the data for each fiscal year
    for i, year in enumerate(fiscal_years):
        for j, category in enumerate(data["Category"]):
            col_letter = increment_letter('B', i)
            if category == "Projection Year" and "E" in year:
                last_projection_year = projection_counter
                data[year][j] = projection_counter
                projection_counter += 1
            elif category == "Unlevered Free Cash Flow":
                data[year][j] = f"='Free Cash Flow'!{col_letter}22"
                last_ufcf_cell = f"{col_letter}{j + 3}"
                
            elif category == "WACC":
                data[year][j] = f"='WACC'!B13" if year == fiscal_years[0] else None
                
            elif category == "Growth Rate":
                data[year][j] = f"=(({last_e_fiscal_year_col_letter}6) / ({last_non_e_fiscal_year_col_letter}3)) ^ (1/{last_e_fiscal_year_col_letter}5) - 1" if year == fiscal_years[0] else None

            elif category == "Enterprise Value":
                data[year][j] = f"= B9 + B13" if year == fiscal_years[0] else None
                
            elif category == "(+) Cash":
                # Set (+) Cash to reference the 'Balance Sheet' sheet, row 4 of the last fiscal year column
                data[year][j] = f"='Balance Sheet'!{last_fiscal_year_col_letter}4" if year == fiscal_years[0] else None
                
            elif category == "(-) Debt":
                data[year][j] = f"='Balance Sheet'!{last_fiscal_year_col_letter}28" if year == fiscal_years[0] else None

            elif category == "(-) Minority Interest":
                data[year][j] = f"='Balance Sheet'!{last_fiscal_year_col_letter}40" if year == fiscal_years[0] else None

            elif category == "Equity Value":
                data[year][j] = f"=B14 + B15 - B16 - B17" if year == fiscal_years[0] else None

            elif category == "Implied Share Price":
                data[year][j] = f"=B18 / B19" if year == fiscal_years[0] else None

            elif category == "Total Shares Outstanding (mm)":
                data[year][j] = total_shares if year == fiscal_years[0] else None

            if category == "Present Value of Free Cash Flow" and "E" in year:
                data[year][j] = f"={col_letter}3 / (1 + B11)^{col_letter}5"
                sum_formula_parts.append(f"{col_letter}6")  # Adjusted to row 6
                last_pvfcf_cell = f"{col_letter}{j + 3}"

    # Construct the SUM formula for "Sum of PV of FCF"
    # Construct the SUM formula for "Sum of PV of FCF" using the updated row references
    if sum_formula_parts:
        sum_formula_index = data["Category"].index("Sum of PV of FCF")
        sum_formula = f"=SUM({','.join(sum_formula_parts)})"
        # Place the SUM formula in the specific cell for the first fiscal year
        first_fiscal_year = fiscal_years[0]
        data[first_fiscal_year][sum_formula_index] = sum_formula

    # Add Terminal Value formula in column 'B' for all fiscal years
    terminal_value_index = data["Category"].index("Terminal Value")
    if last_pvfcf_cell:
        terminal_value_formula = f"=({last_ufcf_cell}) * (1 + B10) / (B11 - B10)"
        for year in fiscal_years:
            data[year][terminal_value_index] = terminal_value_formula if year == fiscal_years[0] else None

    pv_terminal_value_index = data["Category"].index("PV of Terminal Value")
    if last_projection_year:
        pv_terminal_value_formula = f"=B12 * (1 + B11) ^ {last_projection_year}"
        for year in fiscal_years:
            data[year][pv_terminal_value_index] = pv_terminal_value_formula if year == fiscal_years[0] else None
            

    # Convert the data to a DataFrame and set the index to 'Category'
    df = pd.DataFrame(data)
    df.set_index('Category', inplace=True)
    return df
