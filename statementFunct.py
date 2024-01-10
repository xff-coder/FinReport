import freecashflow
import fixed_assets
import statementFunct
import styleModule
import numpy as np
import requests
import json
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle, Border, Side, Color
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE
from openpyxl.utils import get_column_letter
from string import ascii_uppercase


def auto_adjust_column_width(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 5)
        worksheet.column_dimensions[column].width = adjusted_width
        
         
# Function to process, filter, and format data for a given statement
def process_statement(data, columns, selected_columns):
    df = pd.DataFrame(data, columns=columns)
    df.set_index('Fiscal Year', inplace=True)
    df_filtered = df[selected_columns]
    # Convert to millions (without formatting as string)
    for col in df_filtered.columns:
        df_filtered[col] = df_filtered[col] / 1e6
    return df_filtered.T

# Extract Needed Data
def getColumns(shortname):
    if (shortname == 'PL'):
        return ["Revenue",
                
        "Cost of revenue", "Gross Profit",
        "Operating Expenses", "Selling, General & Administrative",
        "Research & Development", "Other Operating Expense",
        "Operating Income (Loss)", "Non-Operating Income (Loss)",
        "Interest Expense, net", "Interest Expense",
        "Interest Income", "Other Investment Income (Loss)",
        "Pretax Income (Loss), Adjusted", "Pretax Income (Loss)",
        "Income Tax (Expense) Benefit, net",
        "Income (Loss) from Affiliates, net of taxes",
        "Income (Loss) from Continuing Operations",
        "Income (Loss) Including Minority Interest",
        "Net Income", "Net Income Available to Common Shareholders"]
    if (shortname == 'BS'):
        return ["Cash, Cash Equivalents & Short Term Investments","Cash & Cash Equivalents","Short Term Investments","Accounts & Notes Receivable",
        "Accounts Receivable, Net","Inventories","Other Short Term Assets", "Prepaid Expenses","Total Current Assets","Property, Plant & Equipment, Net",
        "Other Long Term Assets","Goodwill", "Other Intangible Assets", "Prepaid Expense", "Deferred Tax Assets (Long Term)",
        "Miscellaneous Long Term Assets","Total Noncurrent Assets","Total Assets","Payables & Accruals", "Accounts Payable",
        "Other Payables & Accruals", "Short Term Debt", "Other Short Term Liabilities", "Deferred Revenue (Short Term)",
        "Total Current Liabilities","Long Term Debt", "Other Long Term Liabilities", "Miscellaneous Long Term Liabilities",
        "Total Noncurrent Liabilities","Total Liabilities","Share Capital & Additional Paid-In Capital", "Common Stock",
        "Additional Paid in Capital", "Treasury Stock", "Retained Earnings", "Other Equity",
        "Equity Before Minority Interest","Minority Interest","Total Equity","Total Liabilities & Equity"]
    if (shortname == 'CF'):
        return ["Net Income/Starting Line", "Depreciation & Amortization", "Non-Cash Items",
        "Change in Working Capital", "(Increase) Decrease in Accounts Receivable",
        "(Increase) Decrease in Inventories", "Increase (Decrease) in Accounts Payable",
        "Increase (Decrease) in Other", "Cash from Operating Activities",
        "Change in Fixed Assets & Intangibles", "Disposition of Fixed Assets & Intangibles",
        "Disposition of Fixed Assets", "Acquisition of Fixed Assets & Intangibles",
        "Purchase of Fixed Assets", "Net Change in Long Term Investment",
        "Decrease in Long Term Investment", "Increase in Long Term Investment",
        "Net Cash From Acquisitions & Divestitures", "Other Investing Activities",
        "Cash from Investing Activities", "Cash From (Repayment of) Debt",
        "Cash From (Repayment of) Short Term Debt, net", "Cash From (Repayment of) Long Term Debt, net",
        "Repayments of Long Term Debt", "Cash From Long Term Debt",
        "Cash From (Repurchase of) Equity", "Decrease in Capital Stock",
        "Cash from Financing Activities", "Net Cash Before Disc. Operations and FX",
        "Net Cash Before FX", "Effect of Foreign Exchange Rates", "Net Changes in Cash"]
    return []

# Define a function to apply bold formatting to specific rows
def apply_bold_to_specific_rows(worksheet, statementName):
    if statementName  == 'Income Statement':        
        bold_rows = [
            "Gross Profit", "Operating Income (Loss)",
            "Pretax Income (Loss), Adjusted", "Pretax Income (Loss)",
            "Income (Loss) Including Minority Interest",
            "Net Income", "Net Income Available to Common Shareholders"
            ]
    elif statementName  == 'Balance Sheet':
        bold_rows = [
        "Total Current Assets",
        "Total Noncurrent Assets",
        "Total Assets",
        "Total Current Liabilities",
        "Total Noncurrent Liabilities",
        "Total Liabilities",
        "Total Equity",
        "Total Liabilities & Equity"
        ]
    else: 
        bold_rows = [
        "Cash from Operating Activities",
        "Cash from Investing Activities",
        "Cash from Financing Activities",
        "Net Cash Before Disc. Operations and FX",
        "Net Cash Before FX",
        "Net Changes in Cash"
        ]
        
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row): # Starting from row 2
        if row[0].value in bold_rows:
            for cell in row:
                cell.font = Font(bold=True)

    for idx, cell in enumerate(worksheet['A']):  # Iterate over all cells in Column A
        if cell.value in bold_rows:
            cell.font = Font(bold=True)
            # Apply thick border to the top of the entire row after the bolded row
            if idx < worksheet.max_row:
                for next_row_cell in worksheet[idx+1]:
                    next_row_cell.border = styleModule.thick_top_border
        else:
            cell.font = Font(bold=False)

    # Apply bold formatting only to specific cells in Column A
    for cell in worksheet['A']:  # Iterate over all cells in Column A
        if cell.value in bold_rows:
            cell.font = Font(bold=True)
        else:
            cell.font = Font(bold=False)
            
def getStatementName(shortname):
    if (shortname == 'PL'):
        return 'Income Statement'
    if (shortname == 'BS'):
        return 'Balance Sheet'
    if (shortname == 'CF'):
        return 'Cash Flow'
    
    return ''