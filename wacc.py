import pandas as pd
import yfinance as yf
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
import numpy as np
import nwc
import freecashflow
import fixed_assets
import statementFunct
import styleModule
import requests
import json
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle, Border, Side, Color
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
from string import ascii_uppercase

def increment_letter(letter, increment):
    # Function to increment a letter by a specified number
    if increment == 0:
        return letter
    else:
        alpha_index = ascii_uppercase.index(letter.upper()) + increment
        return ascii_uppercase[alpha_index % 26] + "A" * (alpha_index // 26)
    
def remove_borders_and_gridlines(worksheet):
    # Iterate through all cells in the worksheet
    for row in worksheet.iter_rows():
        for cell in row:
            # Remove border from each cell
            cell.border = Border()

    # Disable gridlines for the worksheet
    worksheet.sheet_view.showGridLines = False
    
def format_rows(worksheet):
    # Define border styles
    thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    medium_border = Border(top=Side(style='medium'), bottom=Side(style='medium'))
    wacc_fill = PatternFill(start_color='fdfd96', end_color='fdfd96', fill_type='solid')

    # Categories to be bolded and have thin border
    bold_and_thin_border_categories = [
        "Cost of Debt", "Weight of Debt", "Cost of Equity", "Weight of Equity", "Debt + Equity"
    ]

    # Category to be bolded and have medium border
    bold_and_medium_border_category = "WACC"

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
        category = row[0].value
        if category in bold_and_thin_border_categories:
            # Apply bold font and thin border
            for cell in row:
                cell.font = Font(bold=True)
                cell.border = thin_border
        elif category == bold_and_medium_border_category:
            # Apply bold font, medium border, and specific background color
            for cell in row:
                cell.font = Font(bold=True)
                cell.border = medium_border
                cell.fill = wacc_fill
        else:
            # Apply regular font with no border
            for cell in row:
                cell.font = Font(bold=False)
                cell.border = Border()


def auto_adjust_column_widths(worksheet):
    for col in worksheet.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass

        # Adjust the column width (if max_length is 0, set a default width)
        adjusted_width = (max_length + 2) if max_length > 0 else 10
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = adjusted_width
        
def apply_percentage_format_to_rows(worksheet):
    # Rows to be formatted as percentages
    percentage_rows = [
        "% of Debt", "Cost of Debt", "Tax Rate",
        "% Equity", "Cost of Equity",
        "Risk Free Rate", "Market Risk Premium","WACC"
    ]

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):  # Starting from row 2
        if row[0].value in percentage_rows:
            for cell in row:
                cell.number_format = '0.00%'
                
def wacc_Currency(worksheet):

    currency_rows = [
        "Total Debt", "Equity Value", "Interest Expense",
        "Total Liabilities", "Debt + Equity"
    ]
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
        if row[0].value in currency_rows:
            for cell in row:
                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE 
    
    last_row = worksheet.max_row + 2
    note_cell = worksheet.cell(row=last_row, column=1)
    note_cell.value = "*$ Expressed in millions"
    note_cell.font = Font(italic=True)
    note_cell.alignment = Alignment(horizontal='left')                

def get_wacc_dataframe(ticker_symbol, start_year, end_year):
    
    stock_ticker = yf.Ticker(ticker_symbol)
    stock_info = stock_ticker.info

    fiscal_years = list(range(start_year + 1, end_year + 1)) + [f"{year}E" for year in range(end_year + 1, end_year + 6)]
    fiscal_years = list(map(str, fiscal_years))
    
    last_fiscal_year_index = len([year for year in fiscal_years if not year.endswith("E")])
    last_fiscal_year_col_letter = increment_letter('B', last_fiscal_year_index)
    
    
    # TNX for risk-free rate
    TNX = yf.Ticker("^TNX")
    beta = stock_info.get("beta", None)
    risk_free_rate = TNX.info['regularMarketPreviousClose'] / 100

    #Market Risk Premium
    # Define the ticker symbol for a broad market index (like the S&P 500)
    sp500 = 0.10


    # Calculate the YTD return
    market_risk_premium = sp500 - risk_free_rate
    
    #Cost of Equity Formula
    cost_of_equity_calculation = f"=B9 + B10 * B11"

    #cost_of_equity_calculation = risk_free_rate + beta * (market_risk_premium)

    # Financials and Balance Sheet
    stock_financials = stock_ticker.financials
    stock_balance_sheet = stock_ticker.balance_sheet

    # Cost of Debt Calculation
    cost_of_debt_calculation = f"=('Income Statement'!{last_fiscal_year_col_letter}12 / B2) * (1 - B5)"

    # Tax Rate (assumed)
    tax_rate = 0.21

    #Equity Value
    market_capitalization = stock_info.get('marketCap', 0) / 1e6  # Convert to millions
    equity_value = market_capitalization

    #Total Debt
    total_debt = f"='Balance Sheet'!{last_fiscal_year_col_letter}24 + 'Balance Sheet'!{last_fiscal_year_col_letter}28"


    #Debt + Equity
    
    debt_equity_addition = f"= B2 + B6"
    
    #% of Debt
    percent_debt = f"= B2 / B12"
    
    #& Equity
    
    percent_equity = f"= B6 / B12"
    
    # WACC Calculation
    wacc_calculation = f"= (B7 * B8) + (B3 * B4 * (1 - B5))"

    # Creating DataFrame
    wacc_data = pd.DataFrame({
        'Category': [
                     "Total Debt",
                     "% of Debt",
                     "Cost of Debt",
                     "Tax Rate",
                     "Equity Value",
                     "% Equity",
                     "Cost of Equity",
                     "Risk Free Rate",
                     "Beta",
                     "Market Risk Premium",
                     "Debt + Equity",
                     "WACC"],
        'Value': [
                  total_debt,
                  percent_debt,
                  cost_of_debt_calculation,
                  tax_rate,
                  equity_value,
                  percent_equity,
                  cost_of_equity_calculation,
                  risk_free_rate,
                  beta,
                  market_risk_premium,
                  debt_equity_addition,
                  wacc_calculation]
    }).set_index('Category')

    return wacc_data