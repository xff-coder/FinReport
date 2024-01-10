import freecashflow
import fixed_assets
import statementFunct
import styleModule
import numpy as np
import requests
import json
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle, Border, Side, Color
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell import MergedCell
from string import ascii_uppercase

# Define the currency format style
currency_style = NamedStyle(name='currency_style')
currency_style.number_format = FORMAT_CURRENCY_USD_SIMPLE
currency_style.alignment = Alignment(horizontal='right')

# Define a style for fiscal year
fiscal_year_style = NamedStyle(name='fiscal_year_style')
fiscal_year_style.font = Font(color=Color(rgb='002596BE'), bold=True) # RGB for the desired blue color
fiscal_year_style.alignment = Alignment(horizontal='center', vertical='center') # Center alignment
# Define a thick top border style
thick_top_border = Border(top=Side(style='thick'))

#define styling for fixed assets
header_style = NamedStyle(name="header_style")
header_style.font = Font(bold=True, size=11)
header_style.alignment = Alignment(horizontal="center", vertical="center")
header_style.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
header_style.border = Border(bottom=Side(border_style="thin"))

data_style = NamedStyle(name="data_style")
data_style.font = Font(size=11)
data_style.alignment = Alignment(horizontal="right", vertical="center")
data_style.number_format = '#,##0.00'

title_style = NamedStyle(name="title_style")
title_style.font = Font(bold=True, size=11, color=Color("FFFFFF"))
title_style.alignment = Alignment(horizontal="center", vertical="center")
title_style.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")

def SetFinStyle(worksheet):
    # Define fill colors for alternating rows
    grey_fill = PatternFill(start_color='00CCCCCC', end_color='00CCCCCC', fill_type='solid')
    white_fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')
    
                    
    # Apply the alternating row colors
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
        fill = grey_fill if row_idx % 2 == 0 else white_fill
        for cell in row:
            cell.fill = fill


    # Set font and alignment for header
    for cell in worksheet['1:1']:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')


    # Set number format for data cells and align column A to the left
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            if cell.column == 1:  # Column A
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE


    # Add a note at the bottom of each sheet
    last_row = worksheet.max_row + 2
    note_cell = worksheet.cell(row=last_row, column=1)
    note_cell.value = "*$ Expressed in millions"
    note_cell.font = Font(italic=True)
    note_cell.alignment = Alignment(horizontal='left')
    
    # Apply the fiscal year style to all cells in the Fiscal Year column
    for cell in worksheet[2]:  # Iterate over all cells in the second row
        if cell.column_letter != 'A':  # Skip the header cell
            cell.value = f'FY {cell.value}'
            cell.style = fiscal_year_style
   




                
    
def SetDCFStyle(worksheet):

    worksheet.sheet_view.showGridLines = False

    # Set font and alignment for header
    for cell in worksheet['1:1']:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left')

    
    # Apply the fiscal year style to all cells in the Fiscal Year column
    for cell in worksheet[2]:  # Iterate over all cells in the second row
        if cell.column_letter != 'A':  # Skip the header cell
            cell.value = f'FY {cell.value}'
            cell.style = fiscal_year_style
            

    # Add a note at the bottom of each sheet
    last_row = worksheet.max_row + 2
    note_cell = worksheet.cell(row=last_row, column=1)
    note_cell.value = "*$ Expressed in millions (Excluding stock price)"
    note_cell.font = Font(italic=True)
    note_cell.alignment = Alignment(horizontal='left')
    
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = Border()

    currency_categories = [
        "Unlevered Free Cash Flow",
        "Present Value of Free Cash Flow",
        "Sum of PV of FCF",
        "Terminal Value",
        "PV of Terminal Value",
        "Enterprise Value",
        "(+) Cash",
        "(-) Debt",
        "(-) Minority Interest",
        "Equity Value",
        "Implied Share Price"
    ]

    # Assuming the categories are in the first column, find their row numbers
    category_rows = {}
    for cell in worksheet['A']:
        if cell.value in currency_categories:
            category_rows[cell.value] = cell.row

    # Apply currency format and right alignment to cells in rows identified above
    for category, row_num in category_rows.items():
        for cell in worksheet[row_num]:
            if cell.column_letter != 'A':  # Apply to all columns except 'A'
                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE

    percentage_categories = ["Growth Rate", "WACC"]

    # Assuming the categories are in the first column, find their row numbers
    for cell in worksheet['A']:
        if cell.value in percentage_categories:
            row_num = cell.row
            for cell in worksheet[row_num]:
                if cell.column_letter != 'A':  # Apply to all columns except 'A'
                    cell.number_format = '0.00%'
               
    categories_group_1 = [
        "Unlevered Free Cash Flow",
        " ",
        "Projection Year",
        "Present Value of Free Cash Flow"
    ]

    categories_group_2 = [
        "Sum of PV of FCF",
        "Growth Rate",
        "WACC",
        "Terminal Value",
        "PV of Terminal Value",
        "Enterprise Value",
        "(+) Cash",
        "(-) Debt",
        "(-) Minority Interest",
        "Equity Value",
        "Total Shares Outstanding (mm)",
        "Implied Share Price"
    ]
     
    all_categories = categories_group_1 + categories_group_2

    for cell in worksheet['A']:
        if cell.value in all_categories:
            row_num = cell.row
            for c in worksheet[row_num]:
                if c.column_letter == 'A':  # Column A: Left-aligned
                    c.alignment = Alignment(horizontal='left')
                else:  # Other columns: Right-aligned with specific formatting
                    c.alignment = Alignment(horizontal='right')
                    if cell.value in currency_categories:
                        c.number_format = FORMAT_CURRENCY_USD_SIMPLE
                    elif cell.value in percentage_categories:
                        c.number_format = '0.00%'

    # Categories Group 1 and 2 - Right-align all except Column A
    for cell in worksheet['A']:
        if cell.value in categories_group_1 + categories_group_2:
            row_num = cell.row
            for c in worksheet[row_num]:
                if c.column_letter != 'A':
                    c.alignment = Alignment(horizontal='right')
            
def SetOtherStyle(worksheet):

    worksheet.sheet_view.showGridLines = False

    # Set font and alignment for header
    for cell in worksheet['1:1']:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left')

    # Set number format for data cells and align column A to the left
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            if cell.column == 1:  # Column A
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
    
    # Apply the fiscal year style to all cells in the Fiscal Year column
    for cell in worksheet[2]:  # Iterate over all cells in the second row
        if cell.column_letter != 'A':  # Skip the header cell
            cell.value = f'FY {cell.value}'
            cell.style = fiscal_year_style
            

    # Add a note at the bottom of each sheet
    last_row = worksheet.max_row + 2
    note_cell = worksheet.cell(row=last_row, column=1)
    note_cell.value = "*$ Expressed in millions"
    note_cell.font = Font(italic=True)
    note_cell.alignment = Alignment(horizontal='left')
    
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = Border()
    
            
def populate_and_style_fiscal_years(worksheet, start_year, end_year, row_number=26):
    fiscal_years = list(range(start_year + 1, end_year + 1)) + [f"{year}E" for year in range(end_year + 1, end_year + 6)]

    for i, year in enumerate(fiscal_years, start=2):  # Assuming Column A is not used for years, start from Column B
        cell = worksheet.cell(row=row_number, column=i)
        cell.value = f"FY {year}"
        cell.style = fiscal_year_style
        
def FApopulate_and_style_fiscal_years(worksheet, start_year, end_year, row_number=9):
    fiscal_years = list(range(start_year + 1, end_year + 1)) + [f"{year}E" for year in range(end_year + 1, end_year + 6)]

    for i, year in enumerate(fiscal_years, start=2):  # Assuming Column A is not used for years, start from Column B
        cell = worksheet.cell(row=row_number, column=i)
        cell.value = f"FY {year}"
        cell.style = fiscal_year_style
        
def NWCpopulate_and_style_fiscal_years(worksheet, start_year, end_year, row_number=18):
    fiscal_years = list(range(start_year, end_year + 1)) + [f"{year}E" for year in range(end_year + 1, end_year + 6)]

    for i, year in enumerate(fiscal_years, start=2):  # Assuming Column A is not used for years, start from Column B
        cell = worksheet.cell(row=row_number, column=i)
        cell.value = f"FY {year}"
        cell.style = fiscal_year_style

def add_border_under_fiscal_year_row(worksheet, row_number):
    # Define a bottom border style
    bottom_border = Border(bottom=Side(style='thin'))
    
    # Apply the border to each cell in the row
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=row_number, column=col)
        cell.border = bottom_border
        
# Define a function to set the background color for data cells
def set_data_cells_background(worksheet, start_row, end_row, start_col, end_col, fill_color):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    if isinstance(start_col, int):
        start_col = get_column_letter(start_col)
    if isinstance(end_col, int):
        end_col = get_column_letter(end_col)

    for row in range(start_row, end_row + 1):
        for col in range(column_index_from_string(start_col), column_index_from_string(end_col) + 1):
            cell = worksheet.cell(row, col)
            cell.fill = fill

def unbold_category_names(worksheet):
    for cell in worksheet['A']:  # Iterate over all cells in Column A
        cell.font = Font(bold=False)