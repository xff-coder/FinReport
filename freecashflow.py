import freecashflow
import fixed_assets
import statementFunct
import styleModule
import numpy as np
import requests
import json
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle, Border, Side, Color
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE
from openpyxl.utils import get_column_letter
from string import ascii_uppercase

def increment_letter(letter, increment):
    # Function to increment a letter by a specified number
    if increment == 0:
        return letter
    else:
        alpha_index = ascii_uppercase.index(letter.upper()) + increment
        return ascii_uppercase[alpha_index % 26] + "A" * (alpha_index // 26)

def style_unlevered_free_cash_flow_row(worksheet, color_hex):
    # Define medium border style
    medium_border = Border(top=Side(style='medium'), bottom=Side(style='medium'))
    fill_color = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

    # Find the row number for "Unlevered Free Cash Flow"
    for row in worksheet.iter_rows(min_row=1, max_col=1):
        for cell in row:
            if cell.value == "Unlevered Free Cash Flow":
                row_number = cell.row
                # Apply a medium border above and below the row, and fill color to all cells in this row
                for col in range(1, worksheet.max_column + 1):
                    cell_to_style = worksheet.cell(row=row_number, column=col)
                    cell_to_style.border = medium_border
                    cell_to_style.fill = fill_color

                    
def auto_adjust_column_width(worksheet, max_width=15, first_col_min_width=15):
    for column in worksheet.columns:
        max_length = 0
        col_letter = column[0].column_letter  # Get the column letter

        for cell in column:
            if cell.value:
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)

        if col_letter == 'A':  # Special handling for the first column
            adjusted_width = max(max_length + 2, first_col_min_width)
        else:
            # Apply a tighter limit for non-column A columns
            adjusted_width = min(max_length, max_width)

        worksheet.column_dimensions[col_letter].width = adjusted_width



categories = [
    "Revenue", "COGS", "Gross Profit", "Operating Expenses",
    "Selling, General, Administrative",
    "Research & Development",
    "Other OpEx",
    "Total Operating Expenses",
    "EBITDA", "Depreciation & Amortization", "Operating Profit (EBIT)",
    "Operating Taxes", "NOPAT (Net Operating Profit After Taxes)",
    "(+) Depreciation & Amortization", "(-) Capital Expenditures",
    "(-) Change in NWC", "NWC", "Current Assets", "Current Liabilities",
    "Unlevered Free Cash Flow",
    "","",
    "Assumptions",
    "Fiscal Year",
    "Revenue Growth",
    "",
    "COGS % of Revenue",
    "Operating Expenses % of Revenue",
    "SG&A % of Revenue",
    "Research & Development % of Revenue",
    "Other OpEx % of Revenue",
    "Tax % of EBIT"
]    
def apply_bold_to_specific_rows(worksheet, statementName):
    if statementName == 'Free Cash Flow':
        bold_rows = [
            "Gross Profit", "EBITDA", "Operating Profit (EBIT)",
            "NOPAT (Net Operating Profit After Taxes)", "Unlevered Free Cash Flow"
        ]
    
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):  # Starting from row 2
        if row[0].value in bold_rows:
            for cell in row:
                cell.font = Font(bold=True)

            # Apply thick border to the top of the entire row after the bolded row
            next_row_idx = row[0].row
            if next_row_idx <= worksheet.max_row:
                for next_row_cell in worksheet[next_row_idx]:
                    next_row_cell.border = Border(top=Side(style='thin'))

    # Apply bold formatting only to specific cells in Column A
    for cell in worksheet['A']:  # Iterate over all cells in Column A
        if cell.value in bold_rows:
            cell.font = Font(bold=True)
        else:
            cell.font = Font(bold=False)
            
            
def apply_percentage_format_to_rows(worksheet):
    # Rows to be formatted as percentages
    percentage_rows = [
        "Revenue Growth", "COGS % of Revenue", "Operating Expenses % of Revenue",
        "SG&A % of Revenue", "Research & Development % of Revenue",
        "Other OpEx % of Revenue", "Tax % of EBIT"
    ]

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):  # Starting from row 2
        if row[0].value in percentage_rows:
            for cell in row:
                cell.number_format = '0.00%'
                
def AddFreeCashFlowSheet(start_year, end_year):
    # Adjust the start year for the free cash flow sheet
    adjusted_start_year = start_year + 1

    # Generate fiscal years starting from adjusted_start_year
    fiscal_years = list(range(adjusted_start_year, end_year + 1)) + [f"{year}E" for year in range(end_year + 1, end_year + 6)]
    fiscal_years = list(map(str, fiscal_years))
    revenue_growth_formulas = []
    cogs_percent_revenue_formulas = []
    sga_percent_opex_formulas = []
    rd_percent_opex_formulas = []
    op_percent_opex_formulas = []
    oop_percent_opex_formulas = []

    data = {"Category": categories}
    
    for i, year in enumerate(fiscal_years):
        is_estimated_year = year.endswith("E")
        data[year] = []
        
        for category in categories:
            col_letter = increment_letter('B', i)
            
            #Revenue Growth
            if category == "Revenue Growth":
                revenue_growth_row_number = categories.index('Revenue Growth') + 3
                
                if i == 0:  # Skip the first fiscal year for revenue growth calculation
                    data[year].append(None)
                    continue
                elif is_estimated_year:
                    # For estimated years, use the average of previous revenue growths
                    if revenue_growth_formulas:
                        avg_formula = f"=AVERAGE({','.join(revenue_growth_formulas)})"
                        data[year].append(avg_formula)
                    else:
                        data[year].append(None)
                    continue
                else:
                    # Calculate revenue growth for non-estimated years
                    prev_year_col_letter = increment_letter('C', i - 2)
                    previous_year_revenue_cell = f"{prev_year_col_letter}3"
                    current_year_revenue_cell = f"{col_letter}3"
                    revenue_growth_formula = f"=({current_year_revenue_cell} - {previous_year_revenue_cell}) / {previous_year_revenue_cell}"
                    revenue_growth_formulas.append(f"{col_letter}{revenue_growth_row_number}")
                    data[year].append(revenue_growth_formula)
                    continue
                

            # Formulas for other categories
            if category == "Revenue":
                if is_estimated_year:
                    prev_year_col_letter = increment_letter('B', i - 1)
                    prev_year_revenue_cell = f"{prev_year_col_letter}3"
                    growth_rate_cell = f"{col_letter}27"
                    formula = f"={prev_year_revenue_cell} * (1 + {growth_rate_cell})"
                else:
                    next_col_letter = increment_letter(col_letter, 1)  # Shift one column ahead
                    formula = f"='Income Statement'!{next_col_letter}3"
            elif category == "COGS":
                if is_estimated_year:
                    current_year_revenue_cell = f"{col_letter}3"
                    cogs_percent_revenue_cell = f"{col_letter}29"
                    formula = f"=-{current_year_revenue_cell} * {cogs_percent_revenue_cell}"
                else:
                    next_col_letter = increment_letter(col_letter, 1)  # Shift one column ahead
                    formula = f"='Income Statement'!{next_col_letter}4"
                    
            elif category == "Gross Profit":
                formula = f"={col_letter}3 + {col_letter}4"
                    
            elif category == "Operating Expenses":
                if is_estimated_year:
                    current_year_revenue_cell = f"{col_letter}3"
                    op_percent_revenue_cell = f"{col_letter}30"
                    formula = f"=-{current_year_revenue_cell} * {op_percent_revenue_cell}"
                else:
                    formula = f"='Income Statement'!{next_col_letter}6"
                
            elif category == "Selling, General, Administrative":
                if is_estimated_year:
                    current_year_revenue_cell = f"{col_letter}3"
                    sga_percent_revenue_cell = f"{col_letter}31"
                    formula = f"=-{current_year_revenue_cell} * {sga_percent_revenue_cell}"
                else:
                    formula = f"='Income Statement'!{next_col_letter}7"
                
            elif category == "Research & Development":
                if is_estimated_year:
                    current_year_revenue_cell = f"{col_letter}3"
                    rd_percent_revenue_cell = f"{col_letter}32"
                    formula = f"=-{current_year_revenue_cell} * {rd_percent_revenue_cell}"
                else:
                    formula = f"='Income Statement'!{next_col_letter}8"
            elif category == "Other OpEx":
                if is_estimated_year:
                    current_year_revenue_cell = f"{col_letter}3"
                    oop_percent_revenue_cell = f"{col_letter}33"
                    formula = f"=-{current_year_revenue_cell} * {oop_percent_revenue_cell}"
                else:
                    formula = f"='Income Statement'!{next_col_letter}9"
            elif category == "Total Operating Expenses":
                formula = f"=SUM({col_letter}6:{col_letter}9)"
            elif category == "EBITDA":
                formula = f"={col_letter}5 + {col_letter}10"
            elif category == "Depreciation & Amortization":
                if is_estimated_year:

                    formula = f"='Fixed Assets'!{col_letter}4"
                else:
                    formula = f"='Cash Flow'!{next_col_letter}4"
            elif category == "Operating Profit (EBIT)":
                formula = f"={col_letter}11 - {col_letter}12"
            elif category == "Operating Taxes":
                formula = '=0.21'
            elif category == "NOPAT (Net Operating Profit After Taxes)":
                formula = f"={col_letter}13 * (1 - {col_letter}14)"
            elif category == "(+) Depreciation & Amortization":
                formula = f"={col_letter}12"
            elif category == "(-) Capital Expenditures":
                formula = f"='Fixed Assets'!{col_letter}5"
            elif category == "COGS % of Revenue":
                revenue_cell = f"{col_letter}3"
                cogs_cell = f"{col_letter}4"
                if is_estimated_year:
                    if cogs_percent_revenue_formulas:
                        avg_formula = f"=AVERAGE({','.join(cogs_percent_revenue_formulas)})"
                        formula = avg_formula
                    else:
                        formula = None
                else:
                    cogs_percent_formula = f"=-{cogs_cell}/{revenue_cell}"
                    cogs_percent_revenue_formulas.append(f"{col_letter}{categories.index('COGS % of Revenue') + 3}")
                    formula = cogs_percent_formula
            elif category == "SG&A % of Revenue":
                revenue_cell = f"{col_letter}3"
                sga_cell = f"{col_letter}7"
                if is_estimated_year:
                    if sga_percent_opex_formulas:
                        avg_formula = f"=AVERAGE({','.join(sga_percent_opex_formulas)})"
                        formula = avg_formula
                    else:
                        formula = None
                else:
                    sga_percent_formula = f"=-{sga_cell}/{revenue_cell}"
                    sga_percent_opex_formulas.append(f"{col_letter}{categories.index('SG&A % of Revenue') + 3}")
                    formula = sga_percent_formula
            elif category == "Research & Development % of Revenue":
                rd_cell = f"{col_letter}8"
                revenue_cell = f"{col_letter}3"
                if is_estimated_year:
                    if rd_percent_opex_formulas:
                        avg_formula = f"=AVERAGE({','.join(rd_percent_opex_formulas)})"
                        formula = avg_formula
                    else:
                        formula = None
                else:
                    rd_percent_formula = f"=-{rd_cell}/{revenue_cell}"
                    rd_percent_opex_formulas.append(f"{col_letter}{categories.index('Research & Development % of Revenue') + 3}")
                    formula = rd_percent_formula
            elif category == "Operating Expenses % of Revenue":
                op_cell = f"{col_letter}6"
                revenue_cell = f"{col_letter}3"
                if is_estimated_year:
                    if op_percent_opex_formulas:
                        avg_formula = f"=AVERAGE({','.join(op_percent_opex_formulas)})"
                        formula = avg_formula
                    else:
                        formula = None
                else:
                    op_percent_formula = f"=-{op_cell}/{revenue_cell}"
                    op_percent_opex_formulas.append(f"{col_letter}{categories.index('Operating Expenses % of Revenue') + 3}")
                    formula = op_percent_formula
            elif category == "Other OpEx % of Revenue":
                oop_cell = f"{col_letter}9"
                revenue_cell = f"{col_letter}3"
                if is_estimated_year:
                    if oop_percent_opex_formulas:
                        avg_formula = f"=AVERAGE({','.join(oop_percent_opex_formulas)})"
                        formula = avg_formula
                    else:
                        formula = None
                else:
                    oop_percent_formula = f"=-{oop_cell}/{revenue_cell}"
                    oop_percent_opex_formulas.append(f"{col_letter}{categories.index('Other OpEx % of Revenue') + 3}")
                    formula = oop_percent_formula
            elif category == "Tax % of EBIT":
                formula = f"={col_letter}14"
            elif category == "Current Assets":
                post_col_letter = increment_letter('B', i + 1)
                formula = f"='Net Working Capital'!{post_col_letter}7"
            elif category == "Current Liabilities":
                post_col_letter = increment_letter('B', i + 1)
                formula = f"='Net Working Capital'!{post_col_letter}13"
            elif category == "NWC":
                formula = f"={col_letter}20 - {col_letter}21"
            elif category == "(-) Change in NWC":
                post_col_letter = increment_letter('B', i + 1)
                formula = f"= {col_letter}19 - 'Net Working Capital'!{col_letter}7 - 'Net Working Capital'!{col_letter}13"
            elif category == "Unlevered Free Cash Flow":
                post_col_letter = increment_letter('B', i + 1)
                formula = f"= {col_letter}15 + {col_letter}16 - {col_letter}17 - {col_letter}18"
            else:    
                formula = ""  # Placeholder for other categories
            data[year].append(formula)
      

                
    df = pd.DataFrame(data)
    df.set_index('Category', inplace=True)
    return df

